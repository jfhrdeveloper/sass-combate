import re, json, unicodedata, collections
from openpyxl import load_workbook

SRC = "/mnt/user-data/uploads/kick1_2026__1_.xlsx"

CLASES = {"A", "B", "C"}
NIVELES = {"AMATEUR", "SEMI PRO", "SEMIPRO", "PRO", "PROFESIONAL"}
MODALIDADES = {
    "LOW KICK": "low_kick", "K1": "k1", "K-1": "k1",
    "KICK LIGHT": "kick_light", "KICKLIGHT": "kick_light",
    "LIGHT CONTACT": "light_contact",
    "POINT FIGHTING": "point_fighting", "POINT FIGTING": "point_fighting",
}


def sin_acentos(s):
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()


def clave(s):
    return re.sub(r"[^a-z0-9]", "", sin_acentos(str(s)).lower())


def limpiar_nombre(s):
    s = re.sub(r"\s+", " ", str(s)).strip()
    seg = bool(re.search(r"segunda\s+pelea", s, re.I))
    s = re.sub(r"[-–]?\s*segunda\s+pelea\s*", "", s, flags=re.I).strip(" -")
    gen = None
    m = re.search(r"\((mujer|hombre|varon|dama)\)", s, re.I)
    if m:
        gen = "F" if m.group(1).lower() in ("mujer", "dama") else "M"
        s = re.sub(r"\((mujer|hombre|varon|dama)\)", "", s, flags=re.I).strip()
    return " ".join(p.capitalize() for p in s.split()), seg, gen


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    t = str(v).strip().replace(",", ".")
    m = re.match(r"^\d+(\.\d+)?$", t)
    return round(float(t), 2) if m else None


def rango_peso(v):
    if v is None:
        return None, None
    if isinstance(v, (int, float)):
        return round(float(v), 2), round(float(v), 2)
    t = str(v).strip().replace(",", ".")
    ns = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", t)]
    if not ns:
        return None, None
    return min(ns), max(ns)


def hora(v):
    if isinstance(v, (int, float)) and 0 < v < 1:
        total = round(v * 24 * 60)
        return f"{total // 60:02d}:{total % 60:02d}"
    if v is None:
        return None
    m = re.search(r"(\d{1,2}):(\d{2})", str(v))
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    if "pm" in str(v).lower() and h < 12:
        h += 12
    return f"{h:02d}:{mi:02d}"


def edades(v):
    if v is None:
        return []
    return [int(x) for x in re.findall(r"\d+", str(v)) if 5 <= int(x) <= 70]


def tiempo(v):
    m = re.search(r"(\d+)\s*x\s*(\d+)\s*x\s*(\d+)", str(v or ""), re.I)
    return (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None


wb = load_workbook(SRC, data_only=True)
clubes_raw = collections.Counter()
crudas = []

for hoja, area in (("RING", "Ring"), ("TATAMI", "Tatami 1")):
    ws = wb[hoja]
    for fila in ws.iter_rows(values_only=True):
        v = list(fila) + [None] * 24
        if not isinstance(v[0], (int, float)) or not v[4] or not v[9]:
            continue
        txt = [str(x).strip() for x in v if x is not None]
        upper = [t.upper() for t in txt]
        clase = next((t for t in upper if t in CLASES), None)
        nivel = next((t.replace("SEMIPRO", "SEMI PRO") for t in upper if t in NIVELES), None)
        modal = next((MODALIDADES[t] for t in upper if t in MODALIDADES), None)
        gan = None
        for i, t in enumerate(upper):
            if t == "GANADOR" and i + 1 < len(txt):
                gan = txt[i + 1]
        estado = "finalizada" if gan else ("lista" if "PROXIMO" in upper else "pendiente")
        na, sa, ga = limpiar_nombre(v[4])
        nb, sb, gb = limpiar_nombre(v[9])
        ta, tb = str(v[5] or "").strip(), str(v[10] or "").strip()
        clubes_raw[ta] += 1
        clubes_raw[tb] += 1
        pmin, pmax = rango_peso(v[3])
        crudas.append(dict(
            area=area, orden=int(v[0]), hora=hora(v[1]),
            sexo=(str(v[2]).strip().upper() if v[2] else None) or ga or gb,
            peso_pactado_min=pmin, peso_pactado_max=pmax,
            roja=dict(nombre=na, club=ta, peso=num(v[6]), segunda=sa),
            azul=dict(nombre=nb, club=tb, peso=num(v[8]), segunda=sb),
            modalidad=modal, edades=edades(v[12]), clase=clase, nivel=nivel,
            formato=tiempo(v[15]) or tiempo(v[16]), estado=estado, ganador_txt=gan,
        ))

canon = {}
for nombre, n in clubes_raw.most_common():
    k = clave(nombre)
    if k and k not in canon:
        canon[k] = " ".join(w.capitalize() for w in re.sub(r"\s+", " ", nombre).strip().split())
clubes = sorted(set(canon.values()))

peleadores = {}
for c in crudas:
    for lado in ("roja", "azul"):
        p = c[lado]
        club = canon.get(clave(p["club"]), "Sin club")
        k = (clave(p["nombre"]), clave(club))
        if k not in peleadores:
            peleadores[k] = dict(nombre=p["nombre"], club=club, pesos=[], edades=[],
                                 sexo=c["sexo"], modalidades=set(), niveles=set(), peleas=0)
        r = peleadores[k]
        r["peleas"] += 1
        if p["peso"]:
            r["pesos"].append(p["peso"])
        r["edades"] += c["edades"]
        if c["modalidad"]:
            r["modalidades"].add(c["modalidad"])
        if c["nivel"]:
            r["niveles"].add(c["nivel"])
        p["club_canon"] = club

lista = []
for i, ((kn, kc), r) in enumerate(sorted(peleadores.items()), 1):
    lista.append(dict(
        id=f"P{i:03d}", nombre=r["nombre"], club=r["club"],
        sexo=r["sexo"], peso=round(sum(r["pesos"]) / len(r["pesos"]), 2) if r["pesos"] else None,
        edad=min(r["edades"]) if r["edades"] else None,
        modalidades=sorted(r["modalidades"]), nivel=sorted(r["niveles"])[0] if r["niveles"] else None,
        peleas_en_evento=r["peleas"],
    ))
idx = {(clave(p["nombre"]), clave(p["club"])): p["id"] for p in lista}

for c in crudas:
    for lado in ("roja", "azul"):
        p = c[lado]
        p["id"] = idx[(clave(p["nombre"]), clave(p["club_canon"]))]
    g = None
    if c["ganador_txt"]:
        kg = clave(limpiar_nombre(c["ganador_txt"])[0])
        for lado in ("roja", "azul"):
            if clave(c[lado]["nombre"]) == kg:
                g = lado
    c["ganador"] = g
    c["ganador_ambiguo"] = bool(c["ganador_txt"]) and g is None

out = dict(evento=dict(nombre="KICK1 Contender Internacional 2026", fecha="2026-07-25",
                       sede="Casa de la Cultura, San Miguel", disciplina="kickboxing"),
           clubes=clubes, peleadores=lista, peleas=crudas)
with open("/home/claude/combate/datos.json", "w") as f:
    json.dump(out, f, ensure_ascii=False, indent=1, default=list)

print(f"variantes de club en el excel : {len(clubes_raw)}")
print(f"clubes reales tras normalizar : {len(clubes)}")
print(f"peleadores unicos             : {len(lista)}")
print(f"peleas                        : {len(crudas)}")
print(f"con 2+ peleas (rompen regla)  : {sum(1 for p in lista if p['peleas_en_evento'] > 1)}")
print(f"ganador no identificable      : {sum(1 for c in crudas if c['ganador_ambiguo'])}")
print(f"sin sexo registrado           : {sum(1 for p in lista if not p['sexo'])}")
print(f"sin edad registrada           : {sum(1 for p in lista if not p['edad'])}")
print(f"sin peso de pesaje            : {sum(1 for p in lista if not p['peso'])}")
